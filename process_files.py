from openpyxl import Workbook, load_workbook


class process_files:

    def process_files(self):
        print("process_files")

        wb = Workbook()
        wb = load_workbook('C:/Users/Albert/Downloads/sample.xlsx')

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 421

        # Save the file
        wb.save("C:/Users/Albert/Downloads/sample.xlsx")

        '''
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        import datetime
        ws['A2'] = datetime.datetime.now()

        # Save the file
        wb.save("C:/Users/Albert/Downloads/sample.xlsx")
        '''